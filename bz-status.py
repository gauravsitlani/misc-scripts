# Authored by Gaurav Sitlani
# Pre-requisite - Create an empty column K Next to Bugzilla number column J

import openpyxl
import requests
import json
from openpyxl import load_workbook

headers = {"Authorization": "Bearer <Bugzilla API Key>"}
wrkbk = load_workbook("<spreadsheet file>.xlsx")

sheet = wrkbk["Sheet1"]

for c in sheet["J"]:
    if type(c.value) is int:
        r = requests.get(
            "https://bugzilla.redhat.com/rest/bug?id=" + str(c.value), headers=headers
        )
        json_data = json.loads(r.text)
        print(str(c.value), json_data["bugs"][0]["status"],
        json_data["bugs"][0]["resolution"])
        sheet.cell(row=c.row, column=11).value = json_data["bugs"][0]["status"]
        sheet.cell(row=c.row, column=12).value = json_data["bugs"][0]["resolution"]

wrkbk.save("<spreadsheet file>.xlsx")
